from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.forms import UserCreationForm, UserChangeForm, PasswordChangeForm
from .utils import render_to_pdf, Calendar
from .models import UserProfile
from django.views import generic
from django.contrib.auth.models import User
from django.http import HttpResponseRedirect
from django.http import HttpResponse
from django.contrib import messages 
from .forms import SignUpForm, EditProfileForm, LoginForm, UserProfileForm, PasswordReset
from .forms import HardWareForm, EventForm
from .models import HardWareForm_table
from django.db.models import Q
import openpyxl
import calendar
from django.utils.safestring import mark_safe
from io import BytesIO
import xlsxwriter
from django.forms.models import model_to_dict
from django.contrib.auth.mixins import LoginRequiredMixin
from datetime import date
#model_to_dict(instance)
import datetime
from .models import *
from django.core.mail import send_mail



class CalendarView(LoginRequiredMixin, generic.ListView):
    login_url = 'signup'
    model = Event
    template_name = 'calendar.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        d = get_date(self.request.GET.get('month', None))
        cal = Calendar(d.year, d.month)
        html_cal = cal.formatmonth(withyear=True)
        context['calendar'] = mark_safe(html_cal)
        context['prev_month'] = prev_month(d)
        context['next_month'] = next_month(d)
        return context

def get_date(req_day):
    if req_day:
        year, month = (int(x) for x in req_day.split('-'))
        return date(year, month, day=1)
    return datetime.today()

def prev_month(d):
    first = d.replace(day=1)
    prev_month = first - timedelta(days=1)
    month = 'month=' + str(prev_month.year) + '-' + str(prev_month.month)
    return month

def next_month(d):
    days_in_month = calendar.monthrange(d.year, d.month)[1]
    last = d.replace(day=days_in_month)
    next_month = last + timedelta(days=1)
    month = 'month=' + str(next_month.year) + '-' + str(next_month.month)
    return month

    '''

     context = super().get_context_data(**kwargs)
        d = get_date(self.request.GET.get('month', None))
        cal = Calendar(d.year, d.month)
        html_cal = cal.formatmonth(withyear=True)
        context['calendar'] = mark_safe(html_cal)
        context['prev_month'] = prev_month(d)
        context['next_month'] = next_month(d)
        return context

    '''


def calendar_test(request):
    msg     = None
    success = False
    context =""
    #context = super().get_context_data(**kwargs)
    d = get_date(request.GET.get('month', None))
    cal = Calendar(d.year, d.month)
    #print(cal)
    html_cal = cal.formatmonth(withyear=True)
    context = mark_safe(html_cal)
    return render(request, "accounts/calendar.html", { "context" : context, "success" : success })


def create_event(request): 
    print("test create event \n\n\n\n")
    form = EventForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        print (request.method)
        event = Event(
                          user=request.user,
                          title = request.POST['title'],
                          description = request.POST['description'],
                          start_date = request.POST['start_date'],
                          end_date = request.POST['end_date'],
                          start_time = request.POST['start_time'],
                          end_time = request.POST['end_time']
                       )
        event.save()
        messages.success(request, ('New Record Added Successfully...'))

        print("test create event ifififififififififi \n\n\n\n")   
        title = request.POST['title']
        description = request.POST['description']
        start_date = request.POST['start_date']
        end_date = request.POST['end_date']
        start_time = request.POST['start_time']
        end_time = request.POST['end_time']
        #calendar_test(request)
        return render(request, "accounts/calendar.html", {})
    else: 
        print("else part")
           
    return render(request, 'accounts/event.html', {'form': form})

class EventEdit(generic.UpdateView):
    model = Event
    fields = ['title', 'description', 'start_time', 'end_time']
    template_name = 'event.html'

#@login_required(login_url='signup')
def event_details(request, event_id):
    event = Event.objects.get(id=event_id)
    eventmember = EventMember.objects.filter(event=event)
    context = {
        'event': event,
        'eventmember': eventmember
    }
    return render(request, 'accounts/event-details.html', context)


def add_eventmember(request, event_id):
    forms = AddMemberForm()
    if request.method == 'POST':
        forms = AddMemberForm(request.POST)
        if forms.is_valid():
            member = EventMember.objects.filter(event=event_id)
            event = Event.objects.get(id=event_id)
            if member.count() <= 9:
                user = forms.cleaned_data['user']
                EventMember.objects.create(
                    event=event,
                    user=user
                )
                return redirect('calendarapp:calendar')
            else:
                print('--------------User limit exceed!-----------------')
    context = {
        'form': forms
    }
    return render(request, 'add_member.html', context)



def asset_search(request):
    submitted = False
    if request.method == 'POST':
        print("export home")
        if request.POST.get("export"):
            print ("export")
        
        context ={}
        AssetNo = request.POST['AssetNo']
        Owner = request.POST['Owner']
        AssetTypeModel = request.POST['AssetTypeModel']
        Group = request.POST['Group']
        TeamName = request.POST['TeamName']
        ProductLine = request.POST['ProductLine']
        Remark = request.POST['Remark']
        # store the data in session 
        request.session['AssetNo']  = request.POST['AssetNo']
        request.session['Owner'] = request.POST['Owner']
        request.session['AssetTypeModel'] = request.POST['AssetTypeModel']
        request.session['Group'] = request.POST['Group']
        request.session['TeamName'] = request.POST['TeamName']
        request.session['ProductLine']  = request.POST['ProductLine']
        request.session['Remark']  = request.POST['Remark']

        all_data = HardWareForm_table.objects.filter(Q(Owner__contains=Owner) &
		                                             Q(AssetNo__contains=AssetNo) &
													 Q(Group__contains=Group) &
													 Q(TeamName__contains=TeamName) &
													 Q(AssetTypeModel__contains=AssetTypeModel) &
													 Q(ProductLine__contains=ProductLine)&
													 Q(Remark__contains=Remark)
													)
        
        context = {"name":asset_search_dispaly , "all":all_data}
        return render(request, 'accounts/search_asset_display.html', context)
    else:
        data_a = HardWareForm_table.objects.all()
        form = HardWareForm()
        if 'submitted' in request.GET:
            submitted = True

    return render(request, 'accounts/search_asset.html', {'form': form, 'submitted': submitted, "data_a":data_a})


def asset_search_dispaly(request):
    context = {"form":HardWareForm}
    form = HardWareForm(request.POST)
    if form.is_valid():
        register = HardWareForm_table(AssetNo = form.cleaned_data['AssetNo'],
                          Owner = form.cleaned_data['Owner'],
                          AssetTypeModel = form.cleaned_data['AssetTypeModel'],
                          Group = form.cleaned_data['Group'],
                          TeamName = form.cleaned_data['TeamName'],
                          ProductLine = form.cleaned_data['ProductLine'],
                          Remark = form.cleaned_data['Remark']
						  )
        all_data = HardWareForm_table.objects.filter(AssetNo=register.AssetNo)
        context = {"name":display , "all":all_data}
        messages.success(request, ('display part'))
    return render(request, 'accounts/search_asset_display.html', context)

def add_asset(request):
	context = {"form":HardWareForm}
	form = HardWareForm(request.POST)
	if form.is_valid():
		
		register = HardWareForm_table(AssetNo = form.cleaned_data['AssetNo'],
                          Owner = form.cleaned_data['Owner'],
                          AssetTypeModel = form.cleaned_data['AssetTypeModel'],
                          Group = form.cleaned_data['Group'],
                          TeamName = form.cleaned_data['TeamName'],
                          ProductLine = form.cleaned_data['ProductLine'],
                          Remark = form.cleaned_data['Remark']
						  )
		register.save()
		messages.success(request, ('New Record Added Successfully...'))
	return render(request, 'accounts/add_asset.html', context)
def WriteToExcel(request,weather_data, town=None):
    AssetNo        = request.session['AssetNo']
    Owner          = request.session['Owner']
    AssetTypeModel = request.session['AssetTypeModel'] 
    Group          = request.session['Group'] 
    TeamName       = request.session['TeamName'] 
    ProductLine    = request.session['ProductLine']  
    Remark    = request.session['Remark']
    all_data1 = HardWareForm_table.objects.filter(Q(Owner__contains=Owner) &
		                                             Q(AssetNo__contains=AssetNo) &
													 Q(Group__contains=Group) &
													 Q(TeamName__contains=TeamName) &
													 Q(AssetTypeModel__contains=AssetTypeModel) &
													 Q(ProductLine__contains=ProductLine)&
													 Q(Remark__contains=Remark)
													)
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
 
    # Here we will adding the code to add data
    worksheet = workbook.add_worksheet()
    # Start from the first cell. Rows and columns are zero indexed.
    row = 0
    col = 0
    worksheet.write(row, col,     "AssetNo")
    worksheet.write(row, col + 1, "Owner")
    worksheet.write(row, col + 2, "AssetTypeModel")
    worksheet.write(row, col + 3, "Group")
    worksheet.write(row, col + 4, "TeamName")
    worksheet.write(row, col + 5, "ProductLine")
    worksheet.write(row, col + 6, "Remark")
    
    row = 1 
    
    for data in all_data1:
        #print(data.AssetNo)
        worksheet.write(row, col,     data.AssetNo)
        worksheet.write(row, col + 1, data.Owner)
        worksheet.write(row, col + 2, data.AssetTypeModel)
        worksheet.write(row, col + 3, data.Group)
        worksheet.write(row, col + 4, data.TeamName)
        worksheet.write(row, col + 5, data.ProductLine)
        worksheet.write(row, col + 6, data.Remark)
        row += 1
    workbook.close()
    xlsx_data = output.getvalue()
    # xlsx_data contains the Excel file
    return xlsx_data

def export_xls(request):
    print("welocme to home export_xls")
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Report.xlsx'
    xlsx_data = WriteToExcel(request,"weather_period", "town")
    response.write(xlsx_data)
    return response
    #return render(request, 'authenticate/display.html')

def export_pdf(request):
    AssetNo        = request.session['AssetNo']
    Owner          = request.session['Owner']
    AssetTypeModel = request.session['AssetTypeModel'] 
    Group          = request.session['Group'] 
    TeamName       = request.session['TeamName'] 
    ProductLine    = request.session['ProductLine']  
    Remark    = request.session['Remark']
    all_data = HardWareForm_table.objects.filter(Q(Owner__contains=Owner) &
		                                             Q(AssetNo__contains=AssetNo) &
													 Q(Group__contains=Group) &
													 Q(TeamName__contains=TeamName) &
													 Q(AssetTypeModel__contains=AssetTypeModel) &
													 Q(ProductLine__contains=ProductLine)&
													 Q(Remark__contains=Remark)
													)
    

    report_name=AssetNo+"_"+Owner+"_"+AssetTypeModel+"_"+Group+"_"+TeamName+"_"+ProductLine

    data = {
             "report":report_name ,"date":datetime.datetime.now(), "all":all_data,
        }
    pdf = render_to_pdf('accounts/pdf.html', data)
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=Report.pdf'
    return response

def import_xls(request):
    print("upload xls")
    if "GET" == request.method:
        return render(request, 'accounts/import.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(excel_file)
        # getting all sheets
        sheets = wb.sheetnames
        # getting a particular sheet
        worksheet = wb["Sheet1"]
        # getting active sheet
        active_sheet = wb.active
        # reading a cell
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))

            excel_data.append(row_data)
        data_valid=0
        data=excel_data[0]
        if ("AssetNo" != data[0]):
            messages.success(request, ('AssetNo not macthed with data base field'))
            data_valid=1
        if ("Owner" != data[1]):
            messages.success(request, ('Owner not macthed with data base field'))
            data_valid=1
        if ("AssetTypeModel" != data[2]):
            messages.success(request, ('AssetTypeModel not macthed with data base field'))
            data_valid=1
        if ("Group" != data[3]):
            messages.success(request, ('Group not macthed with data base field'))
            data_valid=1
        if ("TeamName" != data[4]):
            messages.success(request, ('TeamName not macthed with data base field'))
            data_valid=1
        if ("ProductLine" != data[5]):
            messages.success(request, ('ProductLine not macthed with data base field'))
            data_valid=all_data1 = HardWareForm_table.objects.filter(Q(Owner__contains=Owner) &
		                                             Q(AssetNo__contains=AssetNo) &
													 Q(Group__contains=Group) &
													 Q(TeamName__contains=TeamName) &
													 Q(AssetTypeModel__contains=AssetTypeModel) &
													 Q(ProductLine__contains=ProductLine)&
													 Q(Remark__contains=Remark)
													)
    
        
        if ( data_valid == 0):
            for data in excel_data[1:]:
                register = HardWareForm_table(AssetNo = data[0],
                          Owner = data[1],
                          AssetTypeModel = data[2],
                          Group = data[3],
                          TeamName = data[4],
                          ProductLine = data[5],
                          Remark = data[6]
                          )
                register.save()
        else:
            excel_data=[]
            messages.success(request, ('import failed '))

        return render(request, 'accounts/import.html', {"excel_data":excel_data})

def login_view(request):
    form = LoginForm(request.POST or None)
    msg = None
    if request.method == "POST":

        if form.is_valid():
            username = form.cleaned_data.get("username")
            password = form.cleaned_data.get("password")
            team = form.cleaned_data.get("team")
            group = form.cleaned_data.get("group")
            location = form.cleaned_data.get("location")
            phonenumber = form.cleaned_data.get("phonenumber")
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect("/")
            else:    
                msg = 'Invalid credentials'    
        else:
            msg = 'Error validating the form'    

    return render(request, "accounts/login.html", {"form": form, "msg" : msg})

def register_user(request):
    msg     = None
    success = False
    if request.method == "POST":
        form = SignUpForm(request.POST)
        profile_form = UserProfileForm(request.POST)
        if form.is_valid() and profile_form.is_valid():
            username = form.cleaned_data.get("username")
            email = form.cleaned_data.get("email")
            raw_password = form.cleaned_data.get("password1")
            raw_password1 = form.cleaned_data.get("password2")
            user = authenticate(username=username, password=raw_password)
            user = form.save()
            profile = profile_form.save(commit=False)
            profile.user = user
            profile.save()
            msg     = 'User created - please <a href="/login">login</a>.'
            success = True

        else:
            msg = 'Form is not valid'    
    else:
        form = SignUpForm()
        profile_form = UserProfileForm()


    return render(request, "accounts/adduser.html", {"form": form, "profile_form":profile_form, "msg" : msg, "success" : success })

def edituser(request):
    msg     = None
    success = False
    
    UserProfile1 = UserProfile.objects.all()
    
    User1 = User.objects.all()
    user_data = {}

    for data in User1:
       user_data['username']=data.username
       user_data['email']=data.email

    print(user_data)    
        

    return render(request, "accounts/edituser_main.html", {"user":User1, "UserProfile":UserProfile1})

def update(request,username):

    msg     = None
    success = False
    #User.objects.all()
    User2 = User.objects.get(username=username)
   # UserProfile2 = UserProfile.objects.get(user=username)
    form = SignUpForm(instance=User2)
    #profile_form = UserProfileForm(instance=UserProfile2)
    
    return render(request, "accounts/edituser.html", {"form": form, "msg" : msg, "success" : success })



def password_reset_request(request):
	if request.method == "POST":
		password_reset_form = PasswordResetForm(request.POST)
		if password_reset_form.is_valid():
			data = password_reset_form.cleaned_data['email']
			associated_users = User.objects.filter(Q(email=data))
			if associated_users.exists():
				for user in associated_users:
					subject = "Password Reset Requested"
					email_template_name = "main/password/password_reset_email.txt"
					c = {
					"email":user.email,
					'domain':'127.0.0.1:8000',
					'site_name': 'Website',
					"uid": urlsafe_base64_encode(force_bytes(user.pk)),
					"user": user,
					'token': default_token_generator.make_token(user),
					'protocol': 'http',
					}
					email = render_to_string(email_template_name, c)
					try:
						send_mail(subject, email, 'admin@example.com' , [user.email], fail_silently=False)
					except BadHeaderError:
						return HttpResponse('Invalid header found.')
					return redirect ("/password_reset/done/")
	password_reset_form = PasswordResetForm()
	return render(request=request, template_name="main/password/password_reset.html", context={"password_reset_form":password_reset_form})
